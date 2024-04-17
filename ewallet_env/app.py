import datetime
from flask import Flask, render_template, request, redirect, url_for
import stripe
import openpyxl

from Crypto.Cipher import DES3
from hashlib import md5



# from .my_encryption_module import decrypt_data  # Replace with your encryption logic



app = Flask(__name__)

public_key = "pk_test_6pRNASCoBOKtIshFeQd4XMUh"
stripe.api_key = "sk_test_BQokikJOvBiI2HlWgH4olfQ2"



def chiffrer_DES3(texte_clair, cle):
  """
  Fonction pour chiffrer un texte clair avec DES3.

  Args:
      texte_clair (str): Le texte clair à chiffrer (en chaîne de caractères).
      cle (bytes): La clé de chiffrement DES3 (16 ou 24 octets).

  Returns:
      bytes: Le texte chiffré.

  Raises:
      ValueError: Si la longueur de la clé est incorrecte (16 ou 24 octets).
  """

  # Vérifier la longueur de la clé
  print(len(cle))
  print(type(cle))
  if len(cle) not in [16, 24]:
      raise ValueError("La clé DES3 doit avoir une longueur de 16 ou 24 octets.")

  # Dériver une clé DES plus longue à partir de la clé fournie
  # (facultatif, améliore la sécurité)
  hashed_key = md5(cle).digest()[:16]  # Hashage MD5 pour réduire la clé à 16 octets
  cle_derivee = DES3.new(hashed_key, DES3.MODE_ECB).encrypt(hashed_key)[:24]

  # Création d'un objet chiffreur DES3
  cipher = DES3.new(cle_derivee, DES3.MODE_CBC)  # Mode CBC recommandé pour plus de sécurité

  # Rembourrage du texte clair pour une longueur de bloc multiple de 8
  texte_padde = texte_clair.encode('latin-1') + b'\0' * (8 - len(texte_clair) % 8)

  print("Texte clair paddé:", texte_padde)
  print("Length of text clair paddé:", len(texte_padde))

  # Chiffrement du texte clair
  texte_chiffre = cipher.encrypt(texte_padde)

  return texte_chiffre 

# Exemple d'utilisation
# texte_clair = "Ceci est un message secret à chiffrer."
# texte_clair = "lwanzo9@gmail.com"

 



key = b"UneCleSecrete24Octets!!!"
print("Clé DES3 de 16 octets:", key)

# texte_chiffre = chiffrer_DES3(texte_clair, key)
# print("Texte chiffré:", texte_chiffre)



@app.route('/')
def index():
    return render_template('index.html', public_key = public_key)

@app.route('/thankyou')
def thankyou():
  return render_template('thankyou.html')

@app.route('/payment', methods=['POST'])
def payment():
    
    # CUSTOMER INFO
    customer = stripe.Customer.create(email=request.form['stripeEmail'],
                                      source=request.form['stripeToken'])
  

    
  
    donations = []  # List to store donor data

    donations.append({
        "email": customer.id,
        "amount": 19.99,  # Convert to float if needed
        "timestamp": datetime.datetime.now()  # Use datetime for timestamp
    })

    generate_excel_report(donations)
    return redirect(url_for('thankyou'))



def generate_excel_report(donations):

    wb = openpyxl.Workbook()
    ws = wb.active

    # Add headers
    
    ws.cell(row=1, column=1).value = "Email"
    ws.cell(row=1, column=2).value = "Amount"
    ws.cell(row=1, column=3).value = "Timestamp"

    # Populate data (starting from row 2)
    row_num = 2

    for donation in donations:
        ws.cell(row=row_num, column=1).value = chiffrer_DES3(donation['email'],key)
        ws.cell(row=row_num, column=2).value = donation["amount"]
        ws.cell(row=row_num, column=3).value = donation["timestamp"]
        row_num += 1

    # Save the Excel file
    wb.save("donations_report.xlsx")

if __name__ == '__main__':
    app.run()